"""
reporting.py
============
  * build_excel()        -> .xlsx bytes with Summary and Details sheets
  * discrepancy_email()  -> a detailed Markdown email ready to copy out
"""
from __future__ import annotations

import io
from datetime import datetime

import pandas as pd

import schema
from parsing import clean, to_num, fmt_num


# ═══════════════════════════════════════════════════════════════════
#  Excel
# ═══════════════════════════════════════════════════════════════════
def build_excel(sheets: dict[str, pd.DataFrame]) -> bytes:
    """{'Summary': df, 'Details': df} -> xlsx bytes with styled headers."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for name, df in sheets.items():
            d = (df if df is not None else pd.DataFrame()).copy()
            if d.empty:
                d = pd.DataFrame({"Info": ["No records"]})
            d.to_excel(xw, sheet_name=str(name)[:31], index=False)

        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        head_fill = PatternFill("solid", fgColor="1F3B4D")
        head_font = Font(color="FFFFFF", bold=True, size=10)

        for name, df in sheets.items():
            ws = xw.book[str(name)[:31]]
            for c in ws[1]:
                c.fill = head_fill
                c.font = head_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.freeze_panes = "A2"
            d = df if (df is not None and not df.empty) else pd.DataFrame({"Info": ["No records"]})
            for j, col in enumerate(d.columns, start=1):
                width = max(len(str(col)), *(len(str(v)) for v in d[col].head(200))) if len(d) else len(str(col))
                ws.column_dimensions[get_column_letter(j)].width = min(max(width + 2, 10), 45)
            if len(d):
                ws.auto_filter.ref = ws.dimensions
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════
#  Markdown email
# ═══════════════════════════════════════════════════════════════════
_ICON = {
    schema.M_MISSING: "⛔",
    schema.M_QTY: "🔢",
    schema.M_ITEM: "🏷️",
    schema.M_LOT: "🧾",
    schema.M_ASN: "🔀",
    schema.M_EXTRA: "➕",
}

_HUMAN = {
    schema.M_MISSING: "GRN not done — HU not found in Korber inventory",
    schema.M_QTY: "Quantity mismatch between ASN and inventory",
    schema.M_ITEM: "Item number mismatch",
    schema.M_LOT: "Lot / batch number mismatch",
    schema.M_ASN: "HU received under a different ASN number",
    schema.M_EXTRA: "Extra HU in inventory, not on the ASN document",
}


def _md_table(df: pd.DataFrame, cols: list[str], limit: int = 60) -> str:
    d = df.reindex(columns=cols).fillna("").astype(str).head(limit)
    if d.empty:
        return "_(none)_"
    head = "| " + " | ".join(cols) + " |"
    sep = "| " + " | ".join("---" for _ in cols) + " |"
    body = "\n".join("| " + " | ".join(str(v).replace("|", "/") for v in row) + " |"
                     for row in d.values.tolist())
    extra = f"\n\n_… and {len(df) - limit} more line(s). Full list is in the attached report._" \
        if len(df) > limit else ""
    return "\n".join([head, sep, body]) + extra


def discrepancy_email(summary: pd.DataFrame, disc: pd.DataFrame,
                      *, company="EFL", site="", client="",
                      prepared_by="", to="", cc="",
                      run_id="", inventory_note="", auto=False) -> tuple[str, str]:
    """
    Build the discrepancy email.

    summary : ASN_SUMMARY rows for the affected ASNs
    disc    : DISCREPANCY rows
    auto    : True when generated automatically after an inventory upload
    returns : (subject, markdown_body)
    """
    today = datetime.now().strftime("%d-%b-%Y")
    ts = datetime.now().strftime("%d-%b-%Y %H:%M")
    asns = sorted({clean(a) for a in summary["ASN NO"]}) if not summary.empty else []
    asn_txt = ", ".join(asns[:6]) + (f" (+{len(asns) - 6})" if len(asns) > 6 else "")

    n_disc = len(disc)
    n_high = int((disc["SEVERITY"] == "HIGH").sum()) if n_disc else 0
    subject = (f"[{company}{' / ' + site if site else ''}] ASN vs Korber GRN "
               f"{'Mismatch Alert' if auto else 'Discrepancy Report'} "
               f"— {today} — {n_disc} issue(s) on {len(asns)} ASN")

    L: list[str] = []
    L.append("# ASN vs Korber GRN "
             + ("Mismatch Alert" if auto else "Discrepancy Report"))
    L.append("")
    L.append(f"**To:** {to or '—'}  ")
    L.append(f"**Cc:** {cc or '—'}  ")
    L.append(f"**Subject:** {subject}")
    L.append("")
    L.append("---")
    L.append("")
    L.append("Dear Team,")
    L.append("")
    L.append(
        f"Please find below the reconciliation result between the uploaded **ASN document(s)** "
        f"and the **Korber One inventory** snapshot"
        f"{' (' + inventory_note + ')' if inventory_note else ''}. "
        f"**{n_disc} discrepancy line(s)** were found across **{len(asns)} ASN(s)**"
        f"{f', of which {n_high} are high severity' if n_high else ''}. "
        f"Kindly review and confirm the corrective action."
    )
    L.append("")

    # ── 1. Header ──
    L.append("## 1. Report details")
    L.append("")
    L.append("| Field | Value |")
    L.append("| --- | --- |")
    L.append(f"| Report generated | {ts} |")
    L.append(f"| Prepared by | {prepared_by or '—'} |")
    if auto:
        L.append("| Trigger | Generated automatically on inventory upload |")
    L.append(f"| Site / Warehouse | {site or '—'} |")
    L.append(f"| Client | {client or '—'} |")
    L.append(f"| ASN reference(s) | {asn_txt or '—'} |")
    L.append(f"| Reconciliation run | {run_id or '—'} |")
    L.append(f"| Total discrepancy lines | **{n_disc}** |")
    L.append("")

    # ── 2. ASN summary ──
    L.append("## 2. ASN summary")
    L.append("")
    if summary.empty:
        L.append("_(no ASN records)_")
    else:
        cols = ["ASN NO", "TOTAL LINES", "TOTAL QTY", "MATCHED LINES", "MISSING LINES",
                "MISMATCH LINES", "EXTRA LINES", "RECEIVED QTY", "QTY DIFF", "STATUS"]
        L.append(_md_table(summary, [c for c in cols if c in summary.columns], limit=40))
    L.append("")

    # ── 3. Discrepancy by type ──
    L.append("## 3. Discrepancy breakdown by type")
    L.append("")
    if n_disc == 0:
        L.append("✅ **No discrepancies.** All ASN lines tally with Korber inventory.")
        L.append("")
    else:
        grp = disc.groupby("DISCREPANCY TYPE").agg(
            Lines=("DISC ID", "count"),
            ASN_Qty=("ASN QTY", lambda s: sum(to_num(x) for x in s)),
            Inv_Qty=("INV QTY", lambda s: sum(to_num(x) for x in s)),
        ).reset_index()
        L.append("| Type | Meaning | Lines | ASN Qty | Inventory Qty | Qty diff |")
        L.append("| --- | --- | ---: | ---: | ---: | ---: |")
        for _, r in grp.iterrows():
            t = str(r["DISCREPANCY TYPE"])
            icon = _ICON.get(t, "⚠️")
            mean = _HUMAN.get(t, "Multiple field mismatch on the same line")
            L.append(f"| {icon} {t} | {mean} | {int(r['Lines'])} | "
                     f"{fmt_num(r['ASN_Qty'])} | {fmt_num(r['Inv_Qty'])} | "
                     f"{fmt_num(r['Inv_Qty'] - r['ASN_Qty'])} |")
        L.append("")

        # ── 4. Line level detail per ASN ──
        L.append("## 4. Line level details")
        L.append("")
        for asn, g in disc.groupby(disc["ASN NO"].astype(str)):
            L.append(f"### ASN `{asn}` — {len(g)} issue(s)")
            L.append("")
            cols = ["ASN LINE", "HU ID", "ITEM NUMBER", "LOT NUMBER", "ASN QTY",
                    "INV QTY", "QTY DIFF", "DISCREPANCY TYPE", "SEVERITY", "DETAIL"]
            L.append(_md_table(g, cols))
            L.append("")

    # ── 5. Action ──
    L.append("## 5. Requested action")
    L.append("")
    types = set(disc["DISCREPANCY TYPE"]) if n_disc else set()
    acts = []
    if any(schema.M_MISSING in t for t in types):
        acts.append("**Missing HUs** — please confirm whether the goods were physically received. "
                    "If received, complete the Korber GRN; if not, advise the expected arrival date.")
    if any(schema.M_QTY in t for t in types):
        acts.append("**Quantity mismatches** — please recount the affected HUs and confirm the "
                    "physical quantity, or share the short/excess landing note.")
    if any(schema.M_ITEM in t for t in types):
        acts.append("**Item mismatches** — confirm the correct item number and, if required, "
                    "raise an inventory correction in Korber One.")
    if any(schema.M_LOT in t for t in types):
        acts.append("**Lot mismatches** — confirm the correct lot/batch reference for the affected HUs.")
    if any(schema.M_ASN in t for t in types):
        acts.append("**Wrong ASN** — HUs appear under a different ASN in the system; "
                    "please confirm which ASN the stock belongs to.")
    if any(schema.M_EXTRA in t for t in types):
        acts.append("**Extra HUs** — stock exists in inventory that is not on the ASN document; "
                    "please share the amended ASN or confirm the source of the excess.")
    if not acts:
        acts.append("No action required — the ASN is fully reconciled and ready for AX GRN.")
    for i, a in enumerate(acts, 1):
        L.append(f"{i}. {a}")
    L.append("")
    L.append("Kindly revert with your confirmation so we can proceed with the AX GRN posting.")
    L.append("")
    L.append("Best regards,  ")
    L.append(f"{prepared_by or 'Warehouse Operations'}  ")
    L.append(f"{company}{' — ' + site if site else ''}")
    L.append("")
    L.append("---")
    L.append(f"_Auto-generated by the ASN ↔ GRN Control System on {ts}._")

    return subject, "\n".join(L)
